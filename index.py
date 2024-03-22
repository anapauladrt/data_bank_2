import openpyxl as pd
#criando tabela
tabela=pd.Workbook()
#imprimir tabela
print(tabela.sheetnames)
#criando
tabela.create_sheet('names')#arriscando sorted
names_page=tabela['names']
names_page.append(['names'])
tabela.save('names.xlsx')
#add
while True:
    new=input('enter a name')
    names_page.append([new])
    print(tabela.sheetnames)
    tabela.save('names.xlsx')